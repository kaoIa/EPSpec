from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

XLSX_PATH = r"your address\EPSpec_ An Evidence-Guided, Prior-Retrieval Agent for NIR Band Selection\RAG_Prior knowledge\Data\Functional Group.xlsx"
SOURCE_COLUMN = "Nanometers (nm)"
TARGET_COLUMN = "range_nm"
DELTA = 15.0

def build_range(value, delta: float) -> str:
	if pd.isna(value):
		return ""
	try:
		numeric = float(value)
	except (TypeError, ValueError):
		return str(value)
	start = int(round(numeric - delta))
	end = int(round(numeric + delta))
	return f"{start}-{end}"

def _find_column_index(header_cells, target_name: str) -> int | None:
	for idx, cell in enumerate(header_cells, start=1):
		if cell.value == target_name:
			return idx
	return None

def process_file(input_path: Path) -> Path:
	if not input_path.exists():
		raise FileNotFoundError(f"找不到 Excel 文件：{input_path}")

	workbook = load_workbook(input_path)
	worksheet = workbook.active
	header_cells = list(worksheet[1])
	source_col_idx = _find_column_index(header_cells, SOURCE_COLUMN)
	if source_col_idx is None:
		raise ValueError(f"列 '{SOURCE_COLUMN}' 不存在，请检查 Excel 表头。")
	target_col_idx = _find_column_index(header_cells, TARGET_COLUMN)
	if target_col_idx is None:
		target_col_idx = len(header_cells) + 1
		worksheet.cell(row=1, column=target_col_idx, value=TARGET_COLUMN)

	for row in range(2, worksheet.max_row + 1):
		value = worksheet.cell(row=row, column=source_col_idx).value
		worksheet.cell(row=row, column=target_col_idx).value = build_range(value, DELTA)

	workbook.save(input_path)
	return input_path

def main() -> None:
	result_path = process_file(Path(XLSX_PATH))
	print(f"✅ 区间更新完成：{result_path}")

if __name__ == "__main__":
	main()
